import collections.abc
import datetime
import io
import tempfile
import logging

import flask
import openpyxl
import openpyxl.worksheet
import psycopg2.extensions

from pachu.config import config
from pachu.err import UserError

stdout_logger = logging.getLogger('stdout')
DEFAULT_TRANSFERRED_DELTA = datetime.timedelta(weeks=4)
ALLOWED_TRANSFERRED_DELTA = datetime.timedelta(weeks=52)


def workbook_from_records(*, column_names, records_iterable, filters,
                          query_name):
    assert isinstance(column_names, collections.Iterable)
    assert isinstance(records_iterable, collections.Iterable)
    assert isinstance(filters, collections.Mapping)
    assert isinstance(query_name, str)

    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(query_name)
    offset = (1, 1)
    offset = insert_table(sheet=sheet,
                          columns=['Name', 'Value'],
                          records=filters.items(),
                          table_name='Filters',
                          offset=offset)

    offset = (offset[0] + 1, *offset[1:])

    insert_table(sheet=sheet,
                 columns=column_names,
                 records=records_iterable,
                 table_name=query_name,
                 offset=offset)

    return wb


def insert_table(*, sheet, columns, records, table_name, offset):
    assert isinstance(sheet, openpyxl.worksheet.Worksheet)
    assert isinstance(columns, collections.Iterable)
    assert isinstance(records, collections.Iterable)
    assert isinstance(table_name, str)
    assert isinstance(offset, collections.Container)

    def write_row(sheet, values, offset):
        assert isinstance(sheet, openpyxl.worksheet.Worksheet)
        assert isinstance(values, collections.Iterable)
        assert isinstance(offset, collections.Container)

        row, col = offset

        for col_offset, v in enumerate(values):
            sheet.cell(row=row, column=col + col_offset,
                       value=str(v))

        return (row + 1, col)

    offset = write_row(sheet, [table_name], offset)
    offset = write_row(sheet, columns, offset)
    for row in records:
        offset = write_row(sheet, row, offset)

    return offset


def export_credit_history(
        cursor, *,
        column_names=('active', 'reason', 'plan_id',
                      'transfer_amount', 'transferred_at',
                      'airport_from_id', 'airport_to_id',
                      'date_from', 'date_to', 'user_subscription_id'),
        v,
        api_key,
        fly_from=None,
        fly_to=None,
        date_from=None,
        date_to=None,
        transferred_from=None,
        transferred_to=None,
        status=None,
        transfer_amount=None,
        transfer_amount_operator=None,
        group_by=None
):
    cursor.execute('SET statement_timeout=%(time)s',
                   dict(time=config['api_export_credit_history']['timeout']))
    cursor.execute('SELECT id FROM users WHERE api_key=%s', [api_key])

    user_id = cursor.fetchone()[0]

    # transferred_to = transferred_to or datetime.datetime.now()
    # transferred_from = (transferred_from or
    #                     transferred_to - DEFAULT_TRANSFERRED_DELTA)
    #
    # exceeded_format_msg = 'Transferred date range exceeded {}'.format(
    #     ALLOWED_TRANSFERRED_DELTA)
    # assertUser(transferred_to - transferred_from > ALLOWED_TRANSFERRED_DELTA,
    #            msg=exceeded_format_msg,
    #            code='API_CH_EXCEEDED_TRANSFERRED_DELTA')

    query_params = dict(
        user_id=user_id,
        status=status,
        date_from=date_from,
        date_to=date_to,
        transferred_from=transferred_from,
        transferred_to=transferred_to,
        transfer_amount=transfer_amount,
        fly_from=fly_from,
        fly_to=fly_to,
    )
    filters = dict(
        status_filter='AND users_subscriptions.active=%(status)s' if status else '',
        date_from_filter='AND users_subscriptions.date_from >= %(date_from)s' if date_from else '',
        date_to_filter='AND users_subscriptions.date_to <= %(date_to)s' if date_to else '',
        transferred_at_filter='''
        AND account_transfers.transferred_at BETWEEN 
            (%(transferred_from)s AND %(transferred_to)s)
        ''' if transferred_from and transferred_to else '',
        transfer_amount_filter='''
        AND account_transfers.transfer_amount {transfer_amount_operator} %(transfer_amount)s
        '''.format(
            transfer_amount_operator=transfer_amount_operator) if transfer_amount and transfer_amount_operator else '',
        airport_from_filter='''
        AND (ap_from.name=%(fly_from)s OR ap_from.iata_code=%(fly_from)s)
        ''' if fly_from else '',
        airport_to_filter='''
        AND (ap_to.name=%(fly_to)s OR ap_to.iata_code=%(fly_to)s)
        ''' if fly_to else ''
    )
    select_columns = '''
            active,
            reason,
            subscription_plan_id,
            transfer_amount,
            transferred_at,
            airport_from_id,
            airport_to_id, 
            date_from,
            date_to,
            user_subscr_id
    '''
    group_by_clause = ''
    query = '''
    SELECT {select_columns}
    FROM (
        SELECT 
            users_subscriptions.active,
            'initial tax' AS reason,
            subscription_plan_id,
            transfer_amount,
            transferred_at,
            subscriptions.airport_from_id,
            subscriptions.airport_to_id, 
            date_from,
            date_to,
            users_subscriptions.id AS user_subscr_id
        FROM account_transfers
        JOIN user_subscription_account_transfers AS usat
          ON account_transfers.id=usat.account_transfer_id
        JOIN users_subscriptions 
            ON usat.user_subscription_id=users_subscriptions.id
                {status_filter}
                {date_from_filter}
                {date_to_filter}
        JOIN subscriptions
            ON users_subscriptions.subscription_id=subscriptions.id
        WHERE account_transfers.user_id=%(user_id)s
            {transferred_at_filter}
            {transfer_amount_filter}
        UNION ALL
        SELECT 
            users_subscriptions.active,
            'fetch tax' AS reason,
            subscription_plan_id,
            transfer_amount,
            transferred_at,
            airport_from_id,
            airport_to_id, 
            date_from,
            date_to,
            users_subscriptions.id AS user_subscr_id
        FROM account_transfers
        JOIN subscriptions_fetches_account_transfers AS sfat 
            ON account_transfers.id=sfat.account_transfer_id
        JOIN subscriptions_fetches 
            ON sfat.subscription_fetch_id=subscriptions_fetches.id
        JOIN subscriptions 
            ON subscriptions_fetches.subscription_id=subscriptions.id
        JOIN users_subscriptions 
            ON subscriptions.id=users_subscriptions.subscription_id 
                {status_filter}
                {date_from_filter}
                {date_to_filter}
        WHERE account_transfers.user_id=%(user_id)s
            {transferred_at_filter}
            {transfer_amount_filter}
    ) AS taxes
    JOIN airports AS ap_from
      ON taxes.airport_from_id=ap_from.id
        {airport_from_filter}
    JOIN airports AS ap_to
      ON taxes.airport_to_id=ap_to.id
        {airport_to_filter}
    {group_by_clause}
    ORDER BY 1 DESC
    '''.format(
        **filters,
        group_by_clause=group_by_clause,
        select_columns=select_columns,
    )
    try:
        cursor.execute(query, query_params)
    except psycopg2.extensions.QueryCanceledError:
        stdout_logger.exception('Credit history query timed out.')
        raise UserError(msg='Query took too long',
                        code='API_CH_QUERY_TIMEOUT',
                        userMsg='You are trying to export too much data at once.')

    wb = workbook_from_records(query_name='credit_history',
                               column_names=column_names,
                               records_iterable=cursor.fetchall(),
                               filters=query_params)

    tmp = tempfile.NamedTemporaryFile()
    wb.save(tmp.name)
    tmp.seek(0)
    stream = io.BytesIO(tmp.read())
    filename = '{user_id}-${date}'.format(user_id=user_id,
                                          date=datetime.datetime.now())
    return flask.send_file(
        stream,
        mimetype=config['api_export_credit_history']['xlsx_mime_type'],
        as_attachment=True,
        attachment_filename=filename
    )